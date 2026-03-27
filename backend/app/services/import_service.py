"""엑셀 임포트/익스포트 서비스"""
import io
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.standard import StdCode, StdDomain, StdTerm, StdWord


# ── 엑셀 컬럼 매핑 (실제 엑셀 헤더 → DB 컬럼) ──
WORD_COLUMNS = {
    "논리명": "word_name",
    "물리명": "english_name",
    "물리의미": "english_meaning",
    "속성분류어": "attr_classifier",
    "동의어": "synonyms",
    "설명": "description",
}

DOMAIN_COLUMNS = {
    "도메인그룹명": "domain_group",
    "도메인명": "domain_name",
    "도메인논리명": "domain_code",
    "데이터유형": "data_type",
    "길이": "length",
    "소수점": "decimal_places",
    "설명": "description",
}

TERM_COLUMNS = {
    "논리명": "term_name",
    "물리명": "english_name",
    "영문의미": "english_meaning",
    "도메인논리명": "domain_code",
    "도메인 논리 약어": "term_classifier",
    "도메인그룹": "domain_group",
    "데이터유형": "data_type",
    "길이": "length",
    "소수점": "decimal_places",
    "설명": "description",
}

CODE_COLUMNS = {
    "코드그룹": "code_group",
    "논리명": "code_group_name",
    "물리명": "system_name",
    "코드설명": "description",
    "코드구분": "code_type",
    "길이": "seq_no",
    "코드ID": "code_id",
    "코드값": "code_value",
    "코드값명": "code_value_name",
    "정렬순서": "sort_order",
    "부모코드명": "parent_code_name",
    "부모코드값": "parent_code_value",
    "적용시작일자": "effective_date",
    "적용종료일자": "expiration_date",
}


def _read_excel_rows(file_path: str | Path, column_map: dict, sheet_index: int = 0) -> list[dict]:
    """엑셀 파일에서 행 데이터를 읽어 dict 리스트로 변환"""
    wb = load_workbook(str(file_path), read_only=False, data_only=True)
    ws = wb.worksheets[sheet_index]

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        wb.close()
        return []

    # 헤더 → 컬럼 인덱스 매핑
    col_indices = {}
    for idx, header in enumerate(headers):
        if header and str(header).strip() in column_map:
            col_indices[idx] = column_map[str(header).strip()]

    results = []
    for row in rows_iter:
        record = {}
        for idx, db_col in col_indices.items():
            val = row[idx] if idx < len(row) else None
            if val is not None:
                val = str(val).strip() if isinstance(val, str) else val
            record[db_col] = val
        results.append(record)

    wb.close()
    return results


def _clean_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_date(val):
    """문자열/datetime을 date 객체로 변환"""
    from datetime import date, datetime
    if val is None:
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


async def import_words(db: AsyncSession, file_path: str) -> dict:
    """단어사전 임포트"""
    rows = _read_excel_rows(file_path, WORD_COLUMNS)
    imported, skipped = 0, 0

    for row in rows:
        if not row.get("word_name") or not row.get("english_name"):
            skipped += 1
            continue
        # 문자열 정리
        for key in row:
            if isinstance(row[key], str):
                row[key] = row[key].strip() if row[key] else None
        word = StdWord(**row)
        db.add(word)
        imported += 1

    await db.flush()
    return {"imported": imported, "skipped": skipped, "total": len(rows)}


async def import_domains(db: AsyncSession, file_path: str) -> dict:
    """도메인사전 임포트"""
    rows = _read_excel_rows(file_path, DOMAIN_COLUMNS)
    imported, skipped = 0, 0

    for row in rows:
        if not row.get("domain_name") or not row.get("domain_code"):
            skipped += 1
            continue
        row["length"] = _clean_int(row.get("length"))
        row["decimal_places"] = _clean_int(row.get("decimal_places")) or 0
        row["data_type"] = _clean_str(row.get("data_type")) or "VARCHAR"
        row["domain_group"] = _clean_str(row.get("domain_group")) or ""
        domain = StdDomain(**row)
        db.add(domain)
        imported += 1

    await db.flush()
    return {"imported": imported, "skipped": skipped, "total": len(rows)}


async def import_terms(db: AsyncSession, file_path: str) -> dict:
    """용어사전 임포트 (배치)"""
    rows = _read_excel_rows(file_path, TERM_COLUMNS)
    imported, skipped = 0, 0
    batch = []

    for row in rows:
        if not row.get("term_name") or not row.get("english_name"):
            skipped += 1
            continue
        row["length"] = _clean_int(row.get("length"))
        row["decimal_places"] = _clean_int(row.get("decimal_places")) or 0
        batch.append(StdTerm(**row))
        imported += 1

        if len(batch) >= 5000:
            db.add_all(batch)
            await db.flush()
            batch.clear()

    if batch:
        db.add_all(batch)
        await db.flush()

    return {"imported": imported, "skipped": skipped, "total": len(rows)}


async def import_codes(db: AsyncSession, file_path: str) -> dict:
    """코드사전 임포트 (대량 배치, 다중 시트)"""
    wb = load_workbook(str(file_path), read_only=False, data_only=True)
    total_imported, total_skipped = 0, 0

    for sheet_idx in range(len(wb.worksheets)):
        rows = _read_excel_rows(file_path, CODE_COLUMNS, sheet_index=sheet_idx)
        batch = []

        for row in rows:
            if not row.get("code_group") or not row.get("code_id"):
                total_skipped += 1
                continue
            row["seq_no"] = _clean_int(row.get("seq_no"))
            row["sort_order"] = _clean_int(row.get("sort_order")) or 0
            row["effective_date"] = _clean_date(row.get("effective_date"))
            row["expiration_date"] = _clean_date(row.get("expiration_date"))
            # code_group_name 필수
            row["code_group_name"] = _clean_str(row.get("code_group_name")) or row.get("code_group", "")
            batch.append(StdCode(**row))
            total_imported += 1

            if len(batch) >= 5000:
                db.add_all(batch)
                await db.flush()
                batch.clear()

        if batch:
            db.add_all(batch)
            await db.flush()

    wb.close()
    return {"imported": total_imported, "skipped": total_skipped}


async def export_to_excel(db: AsyncSession, dict_type: str) -> io.BytesIO:
    """사전 데이터를 엑셀로 내보내기"""
    from openpyxl import Workbook
    from sqlalchemy import select

    model_map = {
        "word": (StdWord, WORD_COLUMNS),
        "domain": (StdDomain, DOMAIN_COLUMNS),
        "term": (StdTerm, TERM_COLUMNS),
        "code": (StdCode, CODE_COLUMNS),
    }

    model, col_map = model_map[dict_type]
    reverse_map = {v: k for k, v in col_map.items()}

    result = await db.execute(select(model).where(model.is_deleted == False))
    items = result.scalars().all()

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = dict_type

    db_cols = list(col_map.values())
    headers = [reverse_map.get(c, c) for c in db_cols]
    ws.append(headers)

    for item in items:
        row_data = [getattr(item, col, None) for col in db_cols]
        ws.append(row_data)

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output
