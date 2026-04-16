"""
K-water 표준 사전 4종(단어/도메인/용어/코드) 엑셀 → DB 적재 스크립트.

실행:
  docker cp "docs/K-water 데이터 품질/01. 데이터표준" datahub-portal-backend:/tmp/std_dicts
  docker exec datahub-portal-backend python /app/scripts/import_standard_dicts.py
"""
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

engine = create_engine("postgresql+psycopg://datahub:datahub@postgres:5432/datahub")

BASE_DIR = os.environ.get("STD_DICT_DIR", "/tmp/std_dicts")

FILES = {
    "word": "표준 데이터 조회_K-water데이터표준_단어사전.xlsx",
    "domain": "표준 데이터 조회_K-water데이터표준_도메인사전.xlsx",
    "term": "표준 데이터 조회_K-water데이터표준_용어사전.xlsx",
    "code": "표준 데이터 조회_K-water데이터표준_코드사전.xlsx",
}


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    try:
        s = str(v).strip()
        if not s:
            return None
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _characteristics(personal_info, encryption, scramble):
    parts = []
    if _str(personal_info):
        parts.append(f"개인정보:{_str(personal_info)}")
    if _str(encryption):
        parts.append(f"암호화:{_str(encryption)}")
    if _str(scramble):
        parts.append(f"스크램블:{_str(scramble)}")
    return ", ".join(parts) if parts else None


def iter_data_rows(ws):
    """헤더 1행을 스킵하고 본문 row만 yield. read_only 호환."""
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        yield row


def truncate_tables(s):
    """기존 표준사전 데이터 삭제.

    주의: TRUNCATE CASCADE 는 data_classification 등 FK 참조 테이블까지 비우므로 사용 금지.
    대신 DELETE + 시퀀스 리셋으로 처리한다. data_classification 의 std_*_id FK는
    ON DELETE SET NULL 이므로 참조 컬럼만 NULL 로 바뀌고 분류체계 row 는 유지된다.
    """
    print("[0/4] 기존 표준사전 데이터 삭제 중 (DELETE, 자식 테이블 참조는 NULL 처리됨)...")
    for tbl in ("std_word", "std_domain", "std_term", "std_code"):
        s.execute(text(f"DELETE FROM {tbl}"))
        s.execute(text(f"ALTER SEQUENCE {tbl}_id_seq RESTART WITH 1"))
    s.commit()
    print("  OK std_word, std_domain, std_term, std_code 초기화 완료")


def import_words(s):
    path = os.path.join(BASE_DIR, FILES["word"])
    print(f"[1/4] 단어사전 import: {path}")
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    now = datetime.now()
    batch = []
    seen_keys = set()
    inserted = skipped = 0
    for row in iter_data_rows(ws):
        if not row or not row[0]:
            continue
        word_name = _str(row[0])
        english_name = _str(row[1]) or word_name
        english_meaning = _str(row[2])
        attr_classifier = _str(row[3])
        synonyms = _str(row[4])
        description = _str(row[5])
        if not word_name:
            continue
        key = (word_name, english_name)
        if key in seen_keys:
            skipped += 1
            continue
        seen_keys.add(key)
        batch.append({
            "wn": word_name, "en": english_name, "em": english_meaning,
            "ac": attr_classifier, "sy": synonyms, "d": description,
            "st": "ACTIVE", "ca": now,
        })
        if len(batch) >= 1000:
            _flush_words(s, batch)
            inserted += len(batch)
            batch = []
    if batch:
        _flush_words(s, batch)
        inserted += len(batch)
    s.commit()
    print(f"  OK std_word: {inserted}건 적재 (중복 스킵 {skipped}건)")


def _flush_words(s, batch):
    s.execute(text("""
        INSERT INTO std_word (word_name, english_name, english_meaning, attr_classifier, synonyms, description, status, created_at, updated_at, is_deleted)
        VALUES (:wn, :en, :em, :ac, :sy, :d, :st, :ca, :ca, false)
        ON CONFLICT (word_name, english_name) DO NOTHING
    """), batch)


def import_domains(s):
    path = os.path.join(BASE_DIR, FILES["domain"])
    print(f"[2/4] 도메인사전 import: {path}")
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    now = datetime.now()
    batch = []
    seen = set()
    inserted = skipped = 0
    for row in iter_data_rows(ws):
        if not row or not row[0]:
            continue
        dg = _str(row[0])
        dn = _str(row[1])
        dc = _str(row[2])
        dt = _str(row[3]) or "VARCHAR"
        length = _int(row[4])
        decimal = _int(row[5])
        characteristics = _characteristics(row[6], row[7], row[8])
        desc = _str(row[9])
        if not dc or not dn:
            continue
        if dc in seen:
            skipped += 1
            continue
        seen.add(dc)
        batch.append({
            "dg": dg, "dn": dn, "dc": dc, "dt": dt,
            "l": length, "dp": decimal, "ch": characteristics,
            "d": desc, "st": "ACTIVE", "ca": now,
        })
        if len(batch) >= 500:
            _flush_domains(s, batch)
            inserted += len(batch)
            batch = []
    if batch:
        _flush_domains(s, batch)
        inserted += len(batch)
    s.commit()
    print(f"  OK std_domain: {inserted}건 적재 (중복 스킵 {skipped}건)")


def _flush_domains(s, batch):
    s.execute(text("""
        INSERT INTO std_domain (domain_group, domain_name, domain_code, data_type, length, decimal_places, characteristics, description, status, created_at, updated_at, is_deleted)
        VALUES (:dg, :dn, :dc, :dt, :l, :dp, :ch, :d, :st, :ca, :ca, false)
        ON CONFLICT (domain_code) DO NOTHING
    """), batch)


def import_terms(s):
    path = os.path.join(BASE_DIR, FILES["term"])
    print(f"[3/4] 용어사전 import: {path}")
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    now = datetime.now()
    batch = []
    inserted = 0
    for row in iter_data_rows(ws):
        if not row or not row[0]:
            continue
        term_name = _str(row[0])
        english_name = _str(row[1]) or term_name
        english_meaning = _str(row[2])
        domain_code = _str(row[3])
        term_classifier = _str(row[4])
        domain_group = _str(row[5])
        data_type = _str(row[6])
        length = _int(row[7])
        decimal = _int(row[8])
        characteristics = _characteristics(row[9], row[10], row[11])
        desc = _str(row[12])
        if not term_name:
            continue
        batch.append({
            "tn": term_name, "en": english_name, "em": english_meaning,
            "dc": domain_code, "tc": term_classifier, "dg": domain_group,
            "dt": data_type, "l": length, "dp": decimal,
            "ch": characteristics, "d": desc, "st": "ACTIVE", "ca": now,
        })
        if len(batch) >= 2000:
            _flush_terms(s, batch)
            inserted += len(batch)
            batch = []
            if inserted % 10000 == 0:
                print(f"    ... {inserted}건 진행중")
    if batch:
        _flush_terms(s, batch)
        inserted += len(batch)
    s.commit()
    print(f"  OK std_term: {inserted}건 적재")


def _flush_terms(s, batch):
    s.execute(text("""
        INSERT INTO std_term (term_name, english_name, english_meaning, domain_code, term_classifier, domain_group, data_type, length, decimal_places, characteristics, description, status, created_at, updated_at, is_deleted)
        VALUES (:tn, :en, :em, :dc, :tc, :dg, :dt, :l, :dp, :ch, :d, :st, :ca, :ca, false)
    """), batch)


def import_codes(s):
    path = os.path.join(BASE_DIR, FILES["code"])
    print(f"[4/4] 코드사전 import: {path}")
    wb = load_workbook(path, read_only=False, data_only=True)
    now = datetime.now()
    total_inserted = 0
    for sheet_idx, sn in enumerate(wb.sheetnames, 1):
        ws = wb[sn]
        print(f"  - sheet[{sheet_idx}] {sn}")
        batch = []
        sheet_inserted = 0
        for row in iter_data_rows(ws):
            if not row or not row[0]:
                continue
            code_group = _str(row[0])
            code_group_name = _str(row[1])
            table_name = _str(row[2])
            code_description = _str(row[3])
            code_type = _str(row[4])
            code_id = _str(row[6])
            code_value = _str(row[7])
            code_value_name = _str(row[8])
            sort_order = _int(row[9])
            parent_code_name = _str(row[10])
            parent_code_value = _str(row[11])
            description = _str(row[12])
            effective_date = _date(row[13])
            expiration_date = _date(row[14])
            if not code_group or not code_id:
                continue
            batch.append({
                "cg": code_group, "cgn": code_group_name or code_group,
                "tn": table_name, "cdesc": code_description, "ct": code_type,
                "ci": code_id, "cv": code_value, "cvn": code_value_name,
                "so": sort_order, "pn": parent_code_name, "pv": parent_code_value,
                "d": description, "ed": effective_date, "xd": expiration_date,
                "st": "ACTIVE", "ca": now,
            })
            if len(batch) >= 2000:
                _flush_codes(s, batch)
                sheet_inserted += len(batch)
                batch = []
                if sheet_inserted % 20000 == 0:
                    print(f"    ... {sheet_inserted}건 진행중")
        if batch:
            _flush_codes(s, batch)
            sheet_inserted += len(batch)
        s.commit()
        print(f"    OK {sheet_inserted}건")
        total_inserted += sheet_inserted
    print(f"  OK std_code: 총 {total_inserted}건 적재")


def _flush_codes(s, batch):
    s.execute(text("""
        INSERT INTO std_code (code_group, code_group_name, table_name, description, code_type, code_id, code_value, code_value_name, sort_order, parent_code_name, parent_code_value, effective_date, expiration_date, status, created_at, updated_at, is_deleted)
        VALUES (:cg, :cgn, :tn, :cdesc, :ct, :ci, :cv, :cvn, :so, :pn, :pv, :ed, :xd, :st, :ca, :ca, false)
    """), batch)


def main():
    print("=" * 60)
    print("K-water 표준사전 4종 엑셀 import 시작")
    print("=" * 60)
    if not os.path.exists(BASE_DIR):
        print(f"[ERROR] BASE_DIR 없음: {BASE_DIR}")
        sys.exit(1)
    with Session(engine) as s:
        truncate_tables(s)
        import_words(s)
        import_domains(s)
        import_terms(s)
        import_codes(s)
    print("=" * 60)
    print("OK 전체 표준사전 import 완료")
    print("=" * 60)


if __name__ == "__main__":
    main()
